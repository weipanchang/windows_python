#!/usr/bin/env python
"""
Firefox version: 73.0 (64-bit)
"""
#import xml.etree.ElementTree as ET
#import urllib2
import requests, urllib3, sys
#import requests, sys
import re
from path import Path
import os
import holidays
import shutil
from yahoo_fin import stock_info as si
import yfinance as yf
from openpyxl import load_workbook, Workbook

# from bs4 import BeautifulSoup
# import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import NoSuchElementException
import time
import datetime
from datetime import date
# import sys
# from bs4 import BeautifulSoup as bs
from selenium import webdriver
downloadPath = os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\MSFT_Analysis"
Path(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts").chdir()
eXCEL_File = os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\Stock_2.xlsx"
        
#short_cut_url = "https://finance.yahoo.com/quote/AVGO/history?period1=1249516800&period2=1626307200&interval=1d&filter=history&frequency=1d&includeAdjustedClose=true"
stock = ""
#home_dir = os.path.expanduser( '~' )

class Logger(object):

    def __init__(self):
#        global downloadPath
        global stock
        today = date.today()
        #d1 = today.strftime("%m%d%Y")
        self.terminal = sys.stdout
        self.log = open(downloadPath +"\\Summary_Report_From_Microsoft_"+ today.strftime("%m%d%Y") + ".txt" , "a+")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        #this flush method is needed for python 3 compatibility.
        #this handles the flush command by doing nothing.
        #you might want to specify some extra behavior here.
        pass

class get_data:

    #def __init__(self, stock_name, startDate, endDate, downloadPath):
    def __init__(self, stock_or_fund, msft_ticket):
#        global downloadPath
        global stock
#        stock_name = stock
        print ("")
#        print ("Processing " + self.stock_name.upper() +" stock data")
        self.stock_or_fund = stock_or_fund
        self.delay = 0
        self.currentDateTime = datetime.datetime.now()
        self.date = self.currentDateTime.date()
        
        self.profile = webdriver.FirefoxProfile()
        self.profile.set_preference("browser.download.folderList", 2)
        self.profile.set_preference("browser.download.manager.showWhenStarting", False)
        self.profile.set_preference("browser.download.dir", downloadPath)
        self.profile.set_preference("browser.helperApps.neverAsk.openFile", "text/csv,application/x-msexcel,application/excel,application/x-excel,application/vnd.ms-excel,image/png,image/jpeg,text/html,text/plain,application/msword,application/xml")
        self.profile.set_preference("browser.helperApps.neverAsk.saveToDisk", "text/csv,application/x-msexcel,application/excel,application/x-excel,application/vnd.ms-excel,image/png,image/jpeg,text/html,text/plain,application/msword,application/xml")
        self.profile.set_preference("browser.helperApps.alwaysAsk.force", False)
        self.profile.set_preference("browser.download.manager.alertOnEXEOpen", False)
        self.profile.set_preference("browser.download.manager.focusWhenStarting", False)
        self.profile.set_preference("browser.download.manager.useWindow", False)
        self.profile.set_preference("browser.download.manager.showAlertOnComplete", False)
        self.profile.set_preference("browser.download.manager.closeWhenDone", False)
        self.profile.set_preference("browser.cache.disk.enable", False)
        self.profile.set_preference("browser.cache.memory.enable", False)
        self.profile.set_preference("browser.cache.offline.enable", False)
        self.profile.set_preference("network.http.use-cache", False)
        self.desiredCapabilities = DesiredCapabilities.FIREFOX.copy()
        self.desiredCapabilities['firefox_profile'] = self.profile.encoded
        self.options = Options()
        self.options.add_argument("--headless")

        self.driver = webdriver.Firefox(capabilities=self.desiredCapabilities, options=self.options)

        self.driver.set_page_load_timeout(50)
        self.wait = WebDriverWait(self.driver, 200, poll_frequency=1, ignored_exceptions=[ElementNotVisibleException, ElementNotSelectableException])
#        url = "https://finance.yahoo.com/quote/" + self.stock_name + "?p=" + self.stock_name + "&.tsrc=fin-srch"

        self.url = "https://www.msn.com/en-us/money/watchlist?ocid=winp1taskbar&duration=1M"
        self.url_stock = "https://www.msn.com/en-us/money/watchlist?ocid=winp1taskbar&duration=1M&id="+ msft_ticket
        
        i = 0
        while True:
            try:
                self.driver.get(self.url)
                self.driver.delete_all_cookies()
                self.driver.implicitly_wait(10) # seconds
                if 'watchlist?' in str(self.driver.current_url):
                    break
            except TimeoutException:
                pass

    def quit_driver(self):
        self.driver.quit()
        

 #        if stock[1:] not in ['XAX', 'IXIC', 'DJI', 'GSPC', 'NYA']:
    def get_target_data(self,msft_ticket):
        def check_exists_by_css_selector(css_selector):
            try:
                self.driver.find_element(By.CSS_SELECTOR,css_selector)
            except NoSuchElementException:
                return False
            return True
        def check_exists_by_classname(classname):
            try:
                self.driver.find_element(By.CLASS_NAME,classname)
            except NoSuchElementException:
                return False
            return True
        
        def check_exists_by_xpath(xpath):
            try:
                self.driver.find_element(By.XPATH,xpath)
            except NoSuchElementException:
                return False
            return True
        def check_exists_by_tag(tag_name):
            try:
                self.driver.find_element(By.TAG_NAME,tag_name)
            except NoSuchElementException:
                return False
            return True
        
        self.url_stock = "https://www.msn.com/en-us/money/watchlist?ocid=winp1taskbar&duration=1M&id="+ msft_ticket+"&l3=L3_Earnings"
        self.driver.get(self.url_stock)
        self.driver.implicitly_wait(10)
        print ("Display Earning Page... \n\n")
           
        self.driver.find_element(By.XPATH,'/html/body/div[1]/div[1]/div/div[5]/div[2]/div/div[1]/div/div[3]/div[2]/div/div/button[4]/span').click()

        time.sleep(10)
        
        while True:
            if check_exists_by_xpath('//div[@class= "mainPrice color_red-DS-EntryPoint1-1"]'):
                print ('Current Price:   %s' % (self.driver.find_element("xpath",'//div[@class= "mainPrice color_red-DS-EntryPoint1-1"]').text))
                break
            elif check_exists_by_xpath('//div[@class= "mainPrice color_green-DS-EntryPoint1-1"]'):
                print ('Current Price:   %s' % (self.driver.find_element("xpath",'//div[@class= "mainPrice color_green-DS-EntryPoint1-1"]').text))
                break
            else:
                pass
            
        if check_exists_by_xpath('//div[@class = "price_PreAfter"]'):
            print("After Hours:     %s\n" % (self.driver.find_element("xpath",'//div[@class = "price_PreAfter"]').text))
        
        time.sleep(10)    
        elm_list = self.driver.find_elements(By.XPATH,'//span[@class = "summaryValue-DS-EntryPoint1-2"]')
        target = elm_list[0].text.replace('USD','')
        print( "1y Target Est = %s\n" % (target))
        time.sleep(5)
        print("Recommedation:    %s\n" % (self.driver.find_element("xpath",'//h2[@class="suggestion-DS-EntryPoint1-1"]').text))
        print("Price Volatility: %s\n" % elm_list[1].text)
        
        self.url_stock = "https://www.msn.com/en-us/money/watchlist?ocid=winp1taskbar&duration=1M&id="+ msft_ticket
        self.driver.get(self.url_stock)
        print ("Display Summary Page... \n")
        time.sleep(1)

        elm_list = self.driver.find_elements(By.XPATH,'//div[@class = "factsRowValue-DS-EntryPoint1-1"]')
        previous = elm_list[0].text
        print( "Previous Close = %s\n" % (previous))
        
        time.sleep(1)

    print ('\n' *3)
        
    def update_Excel_Table(self): 
        print ("Updating Spreadsheet Data... \n\n")
        wb = load_workbook(eXCEL_File)
        ws =  wb.active
        i = 3
        while ws['B' + str(i)].value is not None:
            print(ws['A' + str(i)].value, end="   ")
            stock = ws['C' + str(i)].value.rstrip()
            Stock_Fund = ws['B' + str(i)].value.rstrip()
            # print(ws['F' + str(i)].value)
            ws['G'+ str(i)] = self.get_Current_Stock_Price(stock, Stock_Fund)
            print(ws['G'+ str(i)].value)
            i += 1

        wb.save(eXCEL_File)
        self.quit_driver()

    def get_Current_Stock_Price(self, stock, Stock_Fund):
        if Stock_Fund != 'Fund':
            return(float(si.get_live_price(stock)))
        else:
            ticket = yf.Ticker(stock)
            return ticket.info['regularMarketPrice']


def main():
#    global downloadPath
    global stock
    try:
        shutil.rmtree(downloadPath)
    except:
         pass
    time.sleep(2)
    try:
        os.mkdir(downloadPath)
    except:
        pass

    now_time = datetime.datetime.now().time()
    print("\nTime: ", now_time, "\n")
    menu_options = { \
    1: 'Download Historical Data', \
    2: 'Get Target Price from Microsoft', \
    3: 'Download Historycal Data And Summary Date', \
    4: 'Update Excel table with Current Stock Price', \
    5: 'Exit', \
    }

    def print_menu():
        print("Enter your choice: \n\n")
        for key in menu_options.keys():
            print (key, '--', menu_options[key] )
        print ("\n")
        return(int(input()))

    def fetch_Stock_Name(stock_Dictionary):
        stock_fund_names =  [line for line in open("STOCK.txt", "r")]
        for stock_fund_name in stock_fund_names:
            if len(stock_fund_name) < 2 or "IGNOR" in stock_fund_name :
                continue

            stock = re.search(r'(\(\^\w+\))', stock_fund_name)
            if stock is None:
                stock = re.search('\(\w+\)', stock_fund_name)
                msft_ticket = re.search('\[\w+\]', stock_fund_name)

            is_stock =  re.search("ETF|Fund",stock_fund_name)
#            print is_stock
            if is_stock:
                if 'ETF' in stock_fund_name:
                    stock_or_fund =  'ETF'
                else:
                    stock_or_fund = 'Fund'
            else:
                stock_or_fund ='STOCK'
            # print(stock_or_fund)
            stock = stock.group().rstrip().rstrip(')').lstrip('(')
            msft_ticket = msft_ticket.group().rstrip().rstrip(']').lstrip('[')
            stock_Dictionary[stock] = [stock_fund_name.rstrip()[:-9]]
            
            stock_Dictionary[stock].append(stock_or_fund)
            stock_Dictionary[stock].append(msft_ticket)
#            print (stock_Dictionary)
#            os.system("pause")
    def option1():
        get_history_data.get_historical_data()
    
    def option2(msft_ticket):
        get_target_data.get_target_data(msft_ticket)
    
    def option3():
        get_history_data.get_historical_data()
        get_history_data.get_Summary_data()

    def option4():
        get_history_data = get_data('STOCK')
        get_history_data.update_Excel_Table()
        
#     while(True):
# #        print_menu()
#         option = ''
#         try:
#             option = print_menu()
#             break
#         except:
#             print('Wrong input. Please enter a number ...')
    option = 2
    if option == 4:
        wb = load_workbook(eXCEL_File)
        try:
            wb.save(eXCEL_File)
            pass
        except:
            print('Please close the Spreadsheet file, Process Aborted!')
            time.sleep(3)
            # self.quit_driver()
            sys.exit()
        option4()

    elif option < 4:

        fetch_Stock_Name(stock_Dictionary:={})
        for stock in stock_Dictionary.keys():

            sys.stdout = Logger()
            print("\n")
            print (("=") * len("Processing " + stock_Dictionary[stock][0] +" data"))
            print ("Processing " + stock_Dictionary[stock][0] +" data")
            print (("=") * len("Processing " + stock_Dictionary[stock][0] +" data"))

            get_target_data = get_data(stock_Dictionary[stock][1], stock_Dictionary[stock][2])
#            get_history_data = get_data(stock_Dictionary[stock][1])
            if option == 1:
                option1()
            elif option == 2:
                option2(stock_Dictionary[stock][2])
            elif option == 3:
                option3()
            get_target_data.quit_driver()
                
    else:
        print('\nProgram Terminated')
        sys.exit()

    # if option < 5:        
    #     get_history_data.quit_driver()
        
if __name__ == "__main__":
    main()
