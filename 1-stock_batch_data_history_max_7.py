#!/usr/bin/env python
"""
Firefox version: 73.0 (64-bit)
"""
#import xml.etree.ElementTree as ET
#import urllib2
import requests, urllib3, sys
#import requests, sys
import re
from path import Path
import os
import holidays
import shutil
from yahoo_fin import stock_info as si
import yfinance as yf
from openpyxl import load_workbook, Workbook

# from bs4 import BeautifulSoup
# import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import NoSuchElementException
import time
import datetime
from datetime import date
# import sys
# from bs4 import BeautifulSoup as bs
from selenium import webdriver
downloadPath = os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\data"
Path(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts").chdir()
eXCEL_File = os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\Stock_2.xlsx"
        
#short_cut_url = "https://finance.yahoo.com/quote/AVGO/history?period1=1249516800&period2=1626307200&interval=1d&filter=history&frequency=1d&includeAdjustedClose=true"
stock = ""
#home_dir = os.path.expanduser( '~' )

class Logger(object):

    def __init__(self):
#        global downloadPath
        global stock
        today = date.today()
        #d1 = today.strftime("%m%d%Y")
        self.terminal = sys.stdout
        self.log = open(downloadPath +"\\Summary_Report_"+ today.strftime("%m%d%Y") + ".txt" , "a+")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        #this flush method is needed for python 3 compatibility.
        #this handles the flush command by doing nothing.
        #you might want to specify some extra behavior here.
        pass

class get_data:

    #def __init__(self, stock_name, startDate, endDate, downloadPath):
    def __init__(self, stock_or_fund):
#        global downloadPath
        global stock
#        stock_name = stock
        print ("")
#        print ("Processing " + self.stock_name.upper() +" stock data")
        self.stock_or_fund = stock_or_fund
        self.delay = 0
        self.currentDateTime = datetime.datetime.now()
        self.date = self.currentDateTime.date()
        
        self.profile = webdriver.FirefoxProfile()
        self.profile.set_preference("browser.download.folderList", 2)
        self.profile.set_preference("browser.download.manager.showWhenStarting", False)
        self.profile.set_preference("browser.download.dir", downloadPath)
        self.profile.set_preference("browser.helperApps.neverAsk.openFile", "text/csv,application/x-msexcel,application/excel,application/x-excel,application/vnd.ms-excel,image/png,image/jpeg,text/html,text/plain,application/msword,application/xml")
        self.profile.set_preference("browser.helperApps.neverAsk.saveToDisk", "text/csv,application/x-msexcel,application/excel,application/x-excel,application/vnd.ms-excel,image/png,image/jpeg,text/html,text/plain,application/msword,application/xml")
        self.profile.set_preference("browser.helperApps.alwaysAsk.force", False)
        self.profile.set_preference("browser.download.manager.alertOnEXEOpen", False)
        self.profile.set_preference("browser.download.manager.focusWhenStarting", False)
        self.profile.set_preference("browser.download.manager.useWindow", False)
        self.profile.set_preference("browser.download.manager.showAlertOnComplete", False)
        self.profile.set_preference("browser.download.manager.closeWhenDone", False)
        self.profile.set_preference("browser.cache.disk.enable", False)
        self.profile.set_preference("browser.cache.memory.enable", False)
        self.profile.set_preference("browser.cache.offline.enable", False)
        self.profile.set_preference("network.http.use-cache", False)
        self.desiredCapabilities = DesiredCapabilities.FIREFOX.copy()
        self.desiredCapabilities['firefox_profile'] = self.profile.encoded
        self.options = Options()
        self.options.add_argument("--headless")

        self.driver = webdriver.Firefox(capabilities=self.desiredCapabilities, options=self.options)

        self.driver.set_page_load_timeout(50)
        self.wait = WebDriverWait(self.driver, 200, poll_frequency=1, ignored_exceptions=[ElementNotVisibleException, ElementNotSelectableException])
#        url = "https://finance.yahoo.com/quote/" + self.stock_name + "?p=" + self.stock_name + "&.tsrc=fin-srch"

        self.url = "https://finance.yahoo.com"
        self.url_stock = "https://finance.yahoo.com/quote/"+stock.upper()+"?p="+stock.upper()
        i = 0
        while True:
            try:
                self.driver.get(self.url)
                self.driver.delete_all_cookies()
                self.driver.implicitly_wait(10) # seconds
#                time.sleep(self.delay + 1)
#                print ("Yahoo finance Page is loaded")
                if 'finance' in str(self.driver.current_url):
                    break
            except TimeoutException:
                pass

    def quit_driver(self):
        self.driver.quit()
        
    def get_historical_data(self): 
        print ("Retrieving Historical Data ")

        self.ts = datetime.datetime.strptime(str(self.date),"%Y-%m-%d")
#        tuple = element.timetuple()
        self.timestamp = str(int(time.mktime(self.ts.timetuple())))

        self.url_history = "https://finance.yahoo.com/quote/" + stock.upper() + "/history?period1=00&period2=" + self.timestamp +"&interval=1d&filter=history&frequency=1d&includeAdjustedClose=true"
#        url_history = "https://finance.yahoo.com/quote/" + stock + "/history?period1=00&period2=1626480000&interval=1d&filter=history&frequency=1d&includeAdjustedClose=true"
        while True:
            time.sleep(self.delay + 1)
            try:
                time.sleep(self.delay + 1)
                self.driver.get(self.url_history)
                self.driver.implicitly_wait(5)
                if stock.upper() in str(self.driver.current_url):
                    break
            except:
                 print ("Yahoo page slow, will reloop!", end=" ")

#        print ("click at Apply")
        try:
            elm = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//span[text()="Apply"]'))).click()
        except TimeoutException:
            pass

        time.sleep(self.delay + 3)

        a_elm = self.driver.find_element("xpath","//a[@class = 'Fl(end) Mt(3px) Cur(p)']")
#        print ("click at download link")
        a_elm.click()
        time.sleep(self.delay + 3)
#        print ('\n')
        
        weekno = datetime.datetime.today().weekday()
        north_america = holidays.US()

        Day_of_today = self.date.strftime("%m-%d-%Y")

        if weekno < 5 and (Day_of_today not in north_america):
#            print("Saving current stock price.")
            current_time = datetime.datetime.now()
            open_time = current_time.replace(hour=6, minute=20, second=0, microsecond=0)
            while True:
                try:
                    self.driver.get(self.url_stock)
                    self.driver.refresh()
                    self.driver.implicitly_wait(5)
                    if stock.upper() in str(self.driver.current_url):
                        break
                except:
     #               print ("Yahoo slow, will reloop!")
                    pass

#             if (current_time > open_time):
#                 print("Saving current stock price.")
#                 
#                 try:
#                     elm = self.driver.find_element("xpath","//div[@class= 'Fw(b) Fl(end)--m Fz(s) C($primaryColor']").text
# #                    print (elm, end = '\n')
#                 except Exception:
#                     pass
# 
#                 Current_price = self.driver.find_element_by_xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div/span[1]').text
# 
#                 Open = self.driver.find_element_by_xpath('//*[@id="quote-summary"]/div[1]/table/tbody/tr[2]/td[2]/span').text
# #                print("Open =  %.2f" %float(Open))
# 
#                 Range_elm = self.driver.find_element_by_xpath('//*[@id="quote-summary"]/div[1]/table/tbody/tr[5]/td[2]').text
#                 Low, High  = Range_elm.split(' - ')[0], Range_elm.split(' - ')[1]
# #                print ("LOW = %s, HIGH = %s" %(Low, High))
# 
#                 Volume_elm = self.driver.find_element("xpath",'//*[@id="quote-summary"]/div[1]/table/tbody/tr[7]/td[2]/span').text
#                 Volume_elm = Volume_elm.replace(',', '')
# #                print("Volume =  %d" %int(Volume_elm))
#  
#                 with open(downloadPath +"\\" + stock.upper() + '.csv', 'a') as file:
# #                       writer = csv.writer(file)
#                     last_row = ['\n',str(self.ts)[:10], ',', Open, ',', High, ',' , Low, ',', Current_price, ',', Current_price, ',', Volume_elm]
#                     file.writelines(last_row)

 #        if stock[1:] not in ['XAX', 'IXIC', 'DJI', 'GSPC', 'NYA']:
    def get_Summary_data(self):
        def check_exists_by_css_selector(css_selector):
            try:
                self.driver.find_element(By.CSS_SELECTOR,css_selector)
            except NoSuchElementException:
                return False
            return True
        def check_exists_by_classname(classname):
            try:
                self.driver.find_element(By.CLASS_NAME,classname)
#                self.driver.find_element(By.CSS_SELECTOR,classname)
            except NoSuchElementException:
                return False
            return True
        
        def check_exists_by_xpath(xpath):
            try:
                self.driver.find_element(By.XPATH,xpath)
            except NoSuchElementException:
                return False
            return True
        def check_exists_by_tag(tag_name):
            try:
                self.driver.find_element(By.TAG_NAME,tag_name)
            except NoSuchElementException:
                return False
            return True
                
        print ("Display Summary Page... \n\n")

        while True:
            try:
                self.driver.get(self.url_stock)
                self.driver.implicitly_wait(5)
                time.sleep(self.delay + 1)
                print(str(self.driver.current_url))
                if stock.upper() in str(self.driver.current_url):
                    break
            except:
                print ("Yahoo page slow, will reloop!", end=" ")
                pass
        time.sleep(2)

        if check_exists_by_xpath('//img[contains(@class,"Mb(6px) Cur(p)")]'):
            print ("OLD Yahoo Finance Page\n")

#        elif check_exists_by_classname("rapid-noclick-resp opt-in-link"):
        elif check_exists_by_xpath('//*[@class="rapid-noclick-resp opt-in-link"]'):
#        else:
            print ("New Yahoo Finance Page\n")
            
        else:
            print ("Not Detected\n")
                    
        while True:
            try:
                print ('Current Price:   %s' % (self.driver.find_element("xpath",'//*[@id="quote-header-info"]/div[3]/div[1]/div[1]/fin-streamer[1]').text))
                break
            except NoSuchElementException:
                try:
                    print ('Current Price:   %s' % (self.driver.find_element("xpath",'/html/body/div[1]/main/section/section/section/article/section[1]/section/div/section[1]/div[1]/fin-streamer[1]').text))
                    break
                except NoSuchElementException:
                    print ('Current Price:   %s' % (self.driver.find_element("xpath",'//*[@id="quote-header-info"]/div[3]/div[1]/div/span[1]').text))
                    break
         
        if self.stock_or_fund == 'STOCK':
 
            if check_exists_by_xpath("//fin-streamer[contains(@data-field,'postMarketPrice')]"):
                print("After Hours:     %s\n" % (self.driver.find_element("xpath","//fin-streamer[contains(@data-field,'postMarketPrice')]").text))
                
            try:
                elm = self.driver.find_element("xpath","//div[@class= 'Fw(b) Fl(end)--m Fz(s) C($primaryColor']").text
                print (elm, end = '\n')
            except Exception:
                pass

            try:
                Open = self.driver.find_element("xpath",'//*[@id="quote-summary"]/div[1]/table/tbody/tr[2]/td[2]').text
            except:
                Open = self.driver.find_element("xpath",'/html/body/div[1]/main/section/section/section/article/div[2]/ul/li[2]/span[2]/fin-streamer').text
            print("Open =  %.2f" %float(Open.replace(',','')))

            try:
                Range_elm = self.driver.find_element("xpath",'//*[@id="quote-summary"]/div[1]/table/tbody/tr[5]/td[2]').text
            except NoSuchElementException:
                Range_elm = self.driver.find_element("xpath",'/html/body/div[1]/main/section/section/section/article/div[2]/ul/li[5]/span[2]/fin-streamer').text
            Low, High  = Range_elm.split(' - ')[0], Range_elm.split(' - ')[1]
            print ("LOW = %s, HIGH = %s" %(Low, High))

            time.sleep(1)

            try:
                table_elm = self.driver.find_element("xpath",'//*[@id="quote-summary"]/div[2]/table/tbody')
                list_elm = table_elm.find_elements("xpath",'//*/tr[2]')
                time.sleep(0.1)
                for elm in list_elm:
                    if 'Beta (5Y Monthly)' in elm.text:
                        print( elm.text)
                time.sleep(0.1)
                list_elm = table_elm.find_elements("xpath",'//*/tr[8]')
                for elm in list_elm:
                    if '1y Target Est' in elm.text:
                        print ("=====> " + elm.text)
                time.sleep(0.1)
                if check_exists_by_classname('span.Fw\(b\).D\(b\)\-\-mobp.C\(\$negativeColor\)'):
                    print("=====> BEARISH")
                elif check_exists_by_classname('span.Fw\(b\).D\(b\)\-\-mobp.C\(\$positiveColor\)'):
                    print("=====> BULLISH")
                else:
#                check_exists_by_classname('span.Fw\(b\).D\(b\)\-\-mobp.C\(\$cInDiv5-2\)')
                    print ("=====> NEUTRAL")
            
            except:
#                print("\n========New Yahoo Finance Page ==========\n")
                beta = self.driver.find_element(By.XPATH,"/html/body/div[1]/main/section/section/section/article/div[2]/ul/li[10]/span[2]").text
                print( "Beta (5Y Monthly = ", beta)
                target = self.driver.find_element(By.XPATH,"/html/body/div[1]/main/section/section/section/article/div[2]/ul/li[16]/span[2]/fin-streamer").text
                print( "1y Target Est =========> ", target)


            try:
                EPS =  self.driver.find_element("xpath",'/html/body/div[1]/div/div/div[1]/div/div[3]/div[1]/div/div[1]/div/div/div/div[2]/div[2]/table/tbody/tr[4]/td[2]').text
            except NoSuchElementException:
                EPS =  self.driver.find_element("xpath",'/html/body/div[1]/main/section/section/section/article/div[2]/ul/li[12]/span[2]/fin-streamer').text
            
            print ("EPS ( > 1 is better ) = %s" %EPS)

            try:
                PE_Rato = self.driver.find_element("xpath",'//*[@id="quote-summary"]/div[2]/table/tbody/tr[3]/td[2]').text
            except NoSuchElementException:
                PE_Rato = self.driver.find_element("xpath",'/html/body/div[1]/main/section/section/section/article/div[2]/ul/li[11]/span[2]/fin-streamer').text
            print ("PE_Rato ( Smaller is better ) = %s" %PE_Rato)

            if self.stock_or_fund == 'ETF':
                try:
                    print ('Current Price:   %s' % (self.driver.find_element("xpath",'//*[@id="quote-header-info"]/div[3]/div[1]/div[1]/fin-streamer[1]').text))
                except:
                    print ('Current Price:   %s' % (self.driver.find_element("xpath",'//*[@id="quote-header-info"]/div[3]/div[1]/div/span[1]').text))
                   
#                Current_price = self.driver.find_element("xpath",'//*[@id="quote-header-info"]/div[3]/div[1]/div/span[1]').text

                Open = self.driver.find_element("xpath",'//*[@id="quote-summary"]/div[1]/table/tbody/tr[2]/td[2]').text
                print("Open =  %.2f" %float(Open.replace(',','')))
                try:
                    PE_Rato = self.driver.find_element("xpath",'//*[@id="quote-summary"]/div[2]/table/tbody/tr[3]/td[2]/span').text
                except:
                    PE_Rato = self.driver.find_element("xpath",'//*[@id="quote-summary"]/div[2]/table/tbody/tr[3]/td[2]').text
                print ("PE_Rato ( Smaller is better ) = %s" %PE_Rato)

                print (self.driver.find_element("xpath",'//*[@id="quote-summary"]/div[2]/table/tbody/tr[2]/td[1]').text, end ='   ')
                print (self.driver.find_element("xpath",'//*[@id="quote-summary"]/div[2]/table/tbody/tr[2]/td[2]').text)

        print ('\n' *3)
        
    def update_Excel_Table(self): 
        print ("Updating Spreadsheet Data... \n\n")
        wb = load_workbook(eXCEL_File)
        ws =  wb.active
        i = 3
        while ws['B' + str(i)].value is not None:
            print(ws['A' + str(i)].value, end="   ")
            stock = ws['C' + str(i)].value.rstrip()
            Stock_Fund = ws['B' + str(i)].value.rstrip()
            # print(ws['F' + str(i)].value)
            ws['G'+ str(i)] = self.get_Current_Stock_Price(stock, Stock_Fund)
            print(ws['G'+ str(i)].value)
            i += 1

        wb.save(eXCEL_File)
        self.quit_driver()

    def get_Current_Stock_Price(self, stock, Stock_Fund):
        if Stock_Fund != 'Fund':
            return(float(si.get_live_price(stock)))
        else:
            ticket = yf.Ticker(stock)
            return ticket.info['regularMarketPrice']
    #         self.url_stock = "https://finance.yahoo.com/quote/"+stock.upper()+"?p="+stock.upper()
    #         while True:
    #             try:
    #                 self.driver.get(self.url_stock)
    #                 self.driver.implicitly_wait(10)
    #                 time.sleep(self.delay + 1)
    #                 # print(stock, str(self.driver.current_url))
    #                 if stock.upper() in str(self.driver.current_url):
    #                     break
    #             except:
    #                 print ("Yahoo page slow, will reloop!", end=" ")
    #                 pass
    # #        print((self.driver.find_element("xpath",'//*[@id="quote-header-info"]/div[3]/div[1]/div/span[1]').text))
    # #        //*[@id="quote-header-info"]/div[3]/div[1]/div[1]/fin-streamer[1]
    # #        return float((self.driver.find_element_by_xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div/span[1]').text))
    #         return (float(self.driver.find_element_by_xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div[1]/fin-streamer[1]').text))

def main():
#    global downloadPath
    global stock
    try:
        shutil.rmtree(downloadPath)
    except:
         pass
    time.sleep(2)
    try:
        os.mkdir(downloadPath)
    except:
        pass

    now_time = datetime.datetime.now().time()
    print("\nTime: ", now_time, "\n")
    menu_options = { \
    1: 'Download Historical Data', \
    2: 'Get Summary Data', \
    3: 'Download Historycal Data And Summary Date', \
    4: 'Update Excel table with Current Stock Price', \
    5: 'Exit', \
    }

    def print_menu():
        print("Enter your choice: \n\n")
        for key in menu_options.keys():
            print (key, '--', menu_options[key] )
        print ("\n")
        return(int(input()))

    def fetch_Stock_Name(stock_Dictionary):
        stock_fund_names =  [line for line in open("STOCK.txt", "r")]
        for stock_fund_name in stock_fund_names:
            if len(stock_fund_name) < 2 or "IGNOR" in stock_fund_name :
                continue

            stock = re.search(r'(\(\^\w+\))', stock_fund_name)
            if stock is None:
                stock = re.search('\(\w+\)', stock_fund_name)

            is_stock =  re.search("ETF|Fund",stock_fund_name)
#            print is_stock
            if is_stock:
                if 'ETF' in stock_fund_name:
                    stock_or_fund =  'ETF'
                else:
                    stock_or_fund = 'Fund'
            else:
                stock_or_fund ='STOCK'
            # print(stock_or_fund)
            stock = stock.group().rstrip().rstrip(')').lstrip('(')
            stock_Dictionary[stock] = [stock_fund_name.rstrip()]
            stock_Dictionary[stock].append(stock_or_fund)
            # print (stock_Dictionary)

    def option1():
        get_history_data.get_historical_data()
    
    def option2():
        get_history_data.get_Summary_data()
    
    def option3():
        get_history_data.get_historical_data()
        get_history_data.get_Summary_data()

    def option4():
        get_history_data = get_data('STOCK')
        get_history_data.update_Excel_Table()
        
    while(True):
#        print_menu()
        option = ''
        try:
            option = print_menu()
            break
        except:
            print('Wrong input. Please enter a number ...')

    if option == 4:
        wb = load_workbook(eXCEL_File)
        try:
            wb.save(eXCEL_File)
            pass
        except:
            print('Please close the Spreadsheet file, Process Aborted!')
            time.sleep(3)
            # self.quit_driver()
            sys.exit()
        option4()

    elif option < 4:

        fetch_Stock_Name(stock_Dictionary:={})
        for stock in stock_Dictionary.keys():

            sys.stdout = Logger()
            print("\n")
            print (("=") * len("Processing " + stock_Dictionary[stock][0] +" data"))
            print ("Processing " + stock_Dictionary[stock][0] +" data")
            print (("=") * len("Processing " + stock_Dictionary[stock][0] +" data"))

            get_history_data = get_data(stock_Dictionary[stock][1])
    
            if option == 1:
                option1()
            elif option == 2:
                option2()
            elif option == 3:
                option3()
            get_history_data.quit_driver()
                
    else:
        print('\nProgram Terminated')
        sys.exit()

    # if option < 5:        
    #     get_history_data.quit_driver()
        
if __name__ == "__main__":
    main()