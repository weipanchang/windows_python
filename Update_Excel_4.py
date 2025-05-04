#!/usr/bin/env python
import requests
import time
import datetime
from datetime import date
from path import Path
import os
import sys
import holidays
import shutil
from openpyxl import load_workbook, Workbook

dataPath = os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\"                                     
downloadPath = os.path.expanduser( '~' ) + '\Downloads\\'
sTock_download_data_file_name = "Stock.xlsx"
sTock_update_data_file_name = "Stock_1.xlsx"
source = downloadPath + sTock_download_data_file_name
destination = dataPath + sTock_update_data_file_name
delay = 1

def update_Excel_Table(xcl, qUote):
    
    print ("Updating Spreadsheet Data... \n\n")
    wb = load_workbook(xcl)
    ws =  wb.active
    currentDateTime = datetime.datetime.now()
    today = date.today()
    i = 3
    while ws['A' + str(i)].value != "Cash":
        if ws['B' + str(i)].value is not None:
            quote = qUote[ws['C' + str(i)].value]
            ws['G'+ str(i)] = quote
            print(ws['A' + str(i)].value, end="   ")
            print(ws['G'+ str(i)].value)
            print("\n")
        i = i + 1            

    wb.save(xcl)

def get_Current_Stock_Price(xcl):
   today = date.today()
   print(today)
   print ("Reading Spreadsheet Data... \n\n")
   qUote_dict =  dict()
   wb = load_workbook(xcl,data_only=True)
   ws =  wb.active
   currentDateTime = datetime.datetime.now()
   today = date.today()
   i = 4
   while ws['C' + str(i)].value != "CASH":
       if ws['D' + str(i)].value is not None:
           sTock = ws['C' + str(i)].value.rstrip()
           qUote = ws['I' + str(i)].value
           qUote_dict[sTock] = qUote
       i = i + 1
   return (qUote_dict)   

def main():
   qUote_dict = get_Current_Stock_Price(source)
   update_Excel_Table(destination, qUote_dict)

if __name__ == '__main__':
    
    try:
        wb = load_workbook(sTock_download_data_file_name)
    except:
        print("Download File not found!.... Process Aborted")
        time.sleep(3)
        sys.exit()

    main()