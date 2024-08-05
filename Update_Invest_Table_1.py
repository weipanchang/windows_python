#!/usr/bin/env python
import requests
import time
import datetime
from datetime import date
from datetime import timedelta
from path import Path
import os
import sys
import holidays
import shutil
from yahoo_fin import stock_info as si
#from yahoo_fin import stock_info as si
import yfinance as yf
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


#eXCEL_File = "C:\\Users\\William Chang\\Documents\\Python Scripts\\Stock_2.xlsx"
eXCEL_File = os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\Watch_List.xlsx"
delay = 1
flag_equal_value_reset_to_default_color = True# True is default

def update_Excel_Table(xcl):

   wb = load_workbook(xcl)
   try:
       wb.save(xcl)
       pass
   except:
       print('Please close the Spreadsheet file, Process Aborted!')
       time.sleep(3)
       sys.exit()

#   print ("Updating Invest Table... \n\n")
#   wb = load_workbook(xcl)
   ws =  wb.active
   currentDateTime = datetime.datetime.now()
   weekno = datetime.datetime.today().weekday()
   north_america = holidays.US()
   today = date.today()

   print ("Updating Invest Table Data... \n\n")
   wb = load_workbook(xcl)
   ws =  wb.active
   fill_cell1 = PatternFill(patternType='solid', fgColor='FC2C03')  #Red
   fill_cell2 = PatternFill(patternType='solid', fgColor='B4C7DC') 
   fill_cell3 = PatternFill(patternType='solid', fgColor='FFD7D7')  #Pink
   fill_cell4 = PatternFill(patternType='solid', fgColor='FFFF00')  #Yellow
   fill_cell5 = PatternFill(patternType='solid', fgColor='B4C7D7')  #blue
   fill_cell6 = PatternFill(patternType='solid', fgColor='dee6ef')  #Grey
   fill_cell7 = PatternFill(patternType='solid', fgColor='00ff00')  #Green
   fill_cell8 = PatternFill(patternType='solid', fgColor='d9d2e9')  #Purple
      
   i = 9
   while ws['D' + str(i)].value != "INDEX":
      if ws['D' + str(i)].value is not None:
         stock = ws['C' + str(i)].value
         with open(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\data\\Summary_Report__From_Yahoo_" + today.strftime("%m%d%Y")+".txt") as Yahoo, \
             open(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\MSFT_Analysis\\Summary_Report_From_Microsoft_" + today.strftime("%m%d%Y")+".txt") as Microsoft,\
             open(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\Tipranks\\Summary_Report_From_Tipranks_" + today.strftime("%m%d%Y")+".txt") as Tipranks,\
             open(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\Chase\\Summary_Report_From_Chase_" + today.strftime("%m%d%Y")+".txt") as Chase,\
             open(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\Stanley\\Summary_Report_From_Stanley_" + today.strftime("%m%d%Y")+".txt") as Stanley,\
             open(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\Stock_Analysis\\Summary_Report_From_Stock_Analysis_" + today.strftime("%m%d%Y")+".txt") as Stock_Analysis,\
             open(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\WallStreetZen\\Summary_Report_From_WallStreetZen_" + today.strftime("%m%d%Y")+".txt") as WallStreetZen,\
             open(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\Prediction\\Individual_Stock_Report__" + today.strftime("%m%d%Y")+".txt") as Prediction:
            n = 0
            line_from_Yahoo = Yahoo.readline()
            while "("+stock+")" not in line_from_Yahoo:
               line_from_Yahoo = Yahoo.readline()
            print (line_from_Yahoo)

            while True:
               line_from_Yahoo = Yahoo.readline()
               if "1y Target Est" in line_from_Yahoo:
                  target_price = line_from_Yahoo.split()[-1].replace(',', '')
                  print("From Yahoo    ", end="\t\t")
                  print(ws['K' + str(i)].value, end="\t")
                  print (target_price, end = "\t" )
                  if ws['K' + str(i)].value > float(target_price):
                     print("-\t", end="\t")
                     ws['K' + str(i)].fill = fill_cell1
                  elif ws['K' + str(i)].value < float(target_price):
                     print ("+\t", end="\t")
                     ws['K' + str(i)].fill = fill_cell4
                  else: 
                     print (" \t", end="\t")
                     if flag_equal_value_reset_to_default_color:
                        ws['K' + str(i)].fill = fill_cell5
                  ws['K'+ str(i)] = float(target_price.strip(' "'))
#                  print("\n")
               # else:
               #    print("1y Target Est not Found")
               #    break
               if "Volume over Average" in line_from_Yahoo:
                  volume_over_average = line_from_Yahoo.split()[-1]
                  print(ws['O' + str(i)].value, end="\t")
                  print (volume_over_average, end = "\n" )
#                  print("-\n") if ws['K' + str(i)].value > float(target_price) else print ("+\n")                  
                  ws['O'+ str(i)] = float(volume_over_average.strip(' "'))
               if "EOT" in line_from_Yahoo:
                  #print("")
                  break
  
            line_from_Microsoft = Microsoft.readline()   
            while "("+stock+")" not in line_from_Microsoft:
               line_from_Microsoft = Microsoft.readline()
            
            while True:
               line_from_Microsoft = Microsoft.readline()
               if "1y Target Est" in line_from_Microsoft:
                  target_price = line_from_Microsoft.split()[-1].replace(',', '')
                  print("From Microsoft    ", end="\t")
                  print(ws['L' + str(i)].value, end="\t")
                  print (target_price, end = "\t" )
                  if ws['L' + str(i)].value > float(target_price):
                     print("-\t")
                     ws['L' + str(i)].fill = fill_cell1
                  elif ws['L' + str(i)].value < float(target_price):
                     print ("+\t")
                     ws['L' + str(i)].fill = fill_cell4
                  else:
                     print (" \t")
                     if flag_equal_value_reset_to_default_color:
                        ws['L' + str(i)].fill = fill_cell3
                  ws['L'+ str(i)] = float(target_price.strip(' "'))
                  break 

            line_from_Tipranks = Tipranks.readline()   
            while "("+stock+")" not in line_from_Tipranks:
               line_from_Tipranks = Tipranks.readline()
            
            while True:
               line_from_Tipranks = Tipranks.readline()
               if "1y Target Est" in line_from_Tipranks:
                  target_price = line_from_Tipranks.split()[-1].replace(',', '')
                  print("From Tipranks    ", end="\t")
                  print(ws['M' + str(i)].value, end="\t")
                  print (target_price, end = "\t" )
                  if ws['M' + str(i)].value > float(target_price):
                     print("-\t")
                     ws['M' + str(i)].fill = fill_cell1
                  elif ws['M' + str(i)].value < float(target_price):
                     print ("+\t")
                     ws['M' + str(i)].fill = fill_cell4
                  else:
                     print (" \t")
                     target_place = target_price.replace(',', '')
                     if flag_equal_value_reset_to_default_color:
                        ws['M' + str(i)].fill = fill_cell6
                  ws['M'+ str(i)] = float(target_price.strip(' "'))
                  break
            
            line_from_Chase = Chase.readline()   
            while "("+stock+")" not in line_from_Chase:
               line_from_Chase = Chase.readline()
               
            while True:
               line_from_Chase = Chase.readline()
               if "1y Target Est" in line_from_Chase:
                  target_price = line_from_Chase.split()[-1].replace(',', '')
                  print("From Chase    ", end="\t\t")
                  print(ws['N' + str(i)].value, end="\t")
                  print (target_price, end = "\t" )
                  if ws['N' + str(i)].value > float(target_price) and float(target_price) != 0:
                     print("-\n")
                     ws['N' + str(i)].fill = fill_cell1
                  elif ws['N' + str(i)].value < float(target_price):
                     print ("+\n")
                     ws['N' + str(i)].fill = fill_cell4
                  else:
                     print (" \t")
                     target_place = target_price.replace(',', '')
                     if flag_equal_value_reset_to_default_color:
                        ws['N' + str(i)].fill = fill_cell7
                  ws['N'+ str(i)] = float(target_price)
                  break 

            line_from_Stanley = Stanley.readline()   
            while "("+stock+")" not in line_from_Stanley:
               line_from_Stanley = Stanley.readline()
               
            while True:
               line_from_Stanley = Stanley.readline()
               if "1y Target Est" in line_from_Stanley:
                  target_price = line_from_Stanley.split()[-1].replace(',', '')
                  print("From Stanley    ", end="\t")
                  print(ws['Q' + str(i)].value, end="\t")
                  print (target_price, end = "\t" )
                  if ws['Q' + str(i)].value > float(target_price) and float(target_price) != 0:
                     print("-\n")
                     ws['Q' + str(i)].fill = fill_cell1
                  elif ws['Q' + str(i)].value < float(target_price):
                     print ("+\n")
                     ws['Q' + str(i)].fill = fill_cell4
                  else:
                     print (" \t")
                     target_place = target_price.replace(',', '')
                     if flag_equal_value_reset_to_default_color:
                        ws['Q' + str(i)].fill = fill_cell8
                  ws['Q'+ str(i)] = float(target_price.strip(' "'))
                  break 

            line_from_Stock_Analysis = Stock_Analysis.readline()   
            while "("+stock+")" not in line_from_Stock_Analysis:
               line_from_Stock_Analysis = Stock_Analysis.readline()
               
            while True:
               line_from_Stock_Analysis = Stock_Analysis.readline()
               if "1y Target Est" in line_from_Stock_Analysis:
                  target_price = line_from_Stock_Analysis.split()[-1].replace(',', '')
                  print("From Stock_Analysis    ", end="\t")
                  print(ws['R' + str(i)].value, end="\t")
                  print (target_price, end = "\t" )
#                  print (type(ws['R' + str(i)].value), ws['R' + str(i)].value)
                  if ws['R' + str(i)].value > float(target_price) and float(target_price) != 0:
                     print("-\n")
                     ws['R' + str(i)].fill = fill_cell1

                  elif ws['R' + str(i)].value < float(target_price):
                     print ("+\n")
                     ws['R' + str(i)].fill = fill_cell4
                  else:
                     print (" \t")
                     target_place = target_price.replace(',', '')
                     if flag_equal_value_reset_to_default_color:
                        ws['R' + str(i)].fill = fill_cell8
                  ws['R'+ str(i)] = float(target_price.strip(' "'))
                  break 

            line_from_WallStreetZen = WallStreetZen.readline()   
            while "("+stock+")" not in line_from_WallStreetZen:
               line_from_WallStreetZen = WallStreetZen.readline()
               
            while True:
               line_from_WallStreetZen = WallStreetZen.readline()
               if "1y Target Est" in line_from_WallStreetZen:
                  target_price = line_from_WallStreetZen.split()[-1].replace(',', '')
                  print("From WallStreetZen    ", end="\t")
                  print(ws['S' + str(i)].value, end="\t")
                  print (target_price, end = "\t" )
#                  print (type(ws['R' + str(i)].value), ws['R' + str(i)].value)
                  if ws['S' + str(i)].value > float(target_price) and float(target_price) != 0:
                     print("-\n")
                     ws['S' + str(i)].fill = fill_cell1

                  elif ws['S' + str(i)].value < float(target_price):
                     print ("+\n")
                     ws['S' + str(i)].fill = fill_cell4
                  else:
                     print (" \t")
                     target_place = target_price.replace(',', '')
                     if flag_equal_value_reset_to_default_color:
                        ws['S' + str(i)].fill = fill_cell8
                  ws['S'+ str(i)] = float(target_price.strip(' "'))
                  break 

            line_from_Prediction = Prediction.readline()   
            while "[ "+stock+" ]" not in line_from_Prediction:
               line_from_Prediction = Prediction.readline()

            while True:
               line_from_Prediction = Prediction.readline()
               if "Current trend" in line_from_Prediction:
                  Current_trend = line_from_Prediction.split()[4]
                  print("Prediction Trend", end="\t")
                  print(ws['P' + str(i)].value, end="\t")
                  print (Current_trend.replace(",",""), end="\t")
                  if ws['P' + str(i)].value > float(Current_trend.replace(",","")):
                     print("-\n")
                  elif ws['P' + str(i)].value < float(Current_trend.replace(",","")):
                     print("+\n")
                  else:
                     print (" \t", end="")
                  ws['P'+ str(i)] = float(Current_trend.replace(",",""))
                  break
      print("")
      i += 1

   wb.save(xcl)
      
def lines_that_contain(string, fp):
    return [line for line in fp if string in line]
   
def get_Current_Stock_Price(stock, Stock_Fund):
   today = date.today()
   currentDateTime = datetime.datetime.now()
   open_time = currentDateTime.replace(hour=13, minute=00, second=0, microsecond=0)

   if Stock_Fund != 'Fund':
       return(float(si.get_live_price(stock)))
   else:
      if (currentDateTime < open_time):
            today  = today - timedelta(days = 1)
   #        print ("\nMutal Fund disaplayed with Previous Day's Quote")
   #       data = yf.download(stock, start=today)
   #       ticket = yf.Ticker(stock)
      try:
         return float(yf.download(stock, start=today).iloc[0,4])

      except:
         return None

def main():
    update_Excel_Table(eXCEL_File)

if __name__ == '__main__':
    
    # try:
    #     request = requests.get("http://www.yahoo.com", timeout=5)
    #     print("Connected to the Internet")
    # except (requests.ConnectionError, requests.Timeout) as exception:
    #     print("No internet connection.\n\n")
    #     time.sleep(delay + 3)
    #     sys.exit()
#    print(sys.argv[0])
    try:
        wb = load_workbook(eXCEL_File)
    except:
        print("File not found!.... Process Aborted")
        sys.exit()
        
    try:
        wb.save(eXCEL_File)
    except:
        print('Please close the Spreadsheet file, Process Aborted!')
        time.sleep(3)
        sys.exit()
    
    main()