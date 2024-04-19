#!/usr/bin/env python
import requests
import time
import datetime
from datetime import date
from datetime import timedelta
from path import Path
import os
import sys
import holidays
import shutil
from yahoo_fin import stock_info as si
#from yahoo_fin import stock_info as si
import yfinance as yf
from openpyxl import load_workbook, Workbook


#eXCEL_File = "C:\\Users\\William Chang\\Documents\\Python Scripts\\Stock_2.xlsx"
eXCEL_File = os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\Byes.xlsx"
delay = 1

def update_Excel_Table(xcl):

   wb = load_workbook(xcl)
   try:
       wb.save(xcl)
       pass
   except:
       print('Please close the Spreadsheet file, Process Aborted!')
       time.sleep(3)
       sys.exit()

#   print ("Updating Invest Table... \n\n")
#   wb = load_workbook(xcl)
   ws =  wb.active
   currentDateTime = datetime.datetime.now()
   weekno = datetime.datetime.today().weekday()
   north_america = holidays.US()
   today = date.today()

   print ("Updating Byes Table Data... \n\n")
   wb = load_workbook(xcl)
   ws =  wb.active 
   i = 4
   while ws['B' + str(i)].value == "STOCK":
#      if ws['A' + str(i)].value is not None:
      stock = ws['A' + str(i)].value
      with open(os.path.expanduser( '~' ) + "\\Documents\\Python Scripts\\Prediction\\Individual_Stock_Report__" + today.strftime("%m%d%Y")+".txt") as Prediction:
         line_from_Prediction = Prediction.readline()   
         while "[ "+stock+" ]" not in line_from_Prediction:
            line_from_Prediction = Prediction.readline()

         while True:
            line_from_Prediction = Prediction.readline()
            if "Prediction Accuracy Rate:" in line_from_Prediction:
               Prediction_Accuracy_Rate = line_from_Prediction.split()[-2]
               print(stock, end="\t")
               print("Prediction Accuracy Rate:", end="\t")
               print(ws['D' + str(i)].value, end="\t")
               print (Prediction_Accuracy_Rate, end="\t")
               # if float(ws['D' + str(i)].value) > float(Prediction_Accuracy_Rate):
               #    print("-\n")
               # elif ws['D' + str(i)].value < float(Prediction_Accuracy_Rate):
               #    print("+\n")
               # else:
               #    print (" \t", end="")
               ws['D'+ str(i)] = float(Prediction_Accuracy_Rate)
               break               
            
         while True:
            line_from_Prediction = Prediction.readline()
            if "Current trend" in line_from_Prediction:
               Current_trend = line_from_Prediction.split()[4]
               print("Prediction Trend", end="\t")
               print(ws['C' + str(i)].value, end="\t")
               print (Current_trend.replace(",",""), end="\t")
               if ws['C' + str(i)].value > float(Current_trend.replace(",","")):
                  print("-\n")
               elif ws['C' + str(i)].value < float(Current_trend.replace(",","")):
                  print("+\n")
               else:
                  print (" \t", end="")
               ws['C'+ str(i)] = float(Current_trend.replace(",",""))
               break
         print("")
      i += 2

   wb.save(xcl)
   
def lines_that_contain(string, fp):
    return [line for line in fp if string in line]
   
def get_Current_Stock_Price(stock, Stock_Fund):
   today = date.today()
   currentDateTime = datetime.datetime.now()
   open_time = currentDateTime.replace(hour=13, minute=00, second=0, microsecond=0)

   if Stock_Fund != 'Fund':
       return(float(si.get_live_price(stock)))
   else:
      if (currentDateTime < open_time):
            today  = today - timedelta(days = 1)
   #        print ("\nMutal Fund disaplayed with Previous Day's Quote")
   #       data = yf.download(stock, start=today)
   #       ticket = yf.Ticker(stock)
      try:
         return float(yf.download(stock, start=today).iloc[0,4])

      except:
         return None

def main():
    update_Excel_Table(eXCEL_File)

if __name__ == '__main__':
    
    # try:
    #     request = requests.get("http://www.yahoo.com", timeout=5)
    #     print("Connected to the Internet")
    # except (requests.ConnectionError, requests.Timeout) as exception:
    #     print("No internet connection.\n\n")
    #     time.sleep(delay + 3)
    #     sys.exit()
    
    try:
        wb = load_workbook(eXCEL_File)
    except:
        print("File not found!.... Process Aborted")
        sys.exit()
        
    try:
        wb.save(eXCEL_File)
    except:
        print('Please close the Spreadsheet file, Process Aborted!')
        time.sleep(3)
        sys.exit()
    
    main()