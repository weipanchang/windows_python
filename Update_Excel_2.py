#!/usr/bin/env python
import requests
import time
from datetime import date
from path import Path
import os
import sys
import holidays
import shutil
from yahoo_fin import stock_info as si
#from yahoo_fin import stock_info as si
import yfinance as yf
from openpyxl import load_workbook, Workbook


#eXCEL_File = "C:\\Users\\William Chang\\Documents\\Python Scripts\\Stock_2.xlsx"
eXCEL_File = ".\\Stock_1.xlsx"
delay = 1

def update_Excel_Table(xcl):
    
    print ("Updating Spreadsheet Data... \n\n")
    wb = load_workbook(xcl)
    ws =  wb.active
    i = 3
    while ws['A' + str(i)].value != "Cash":
        if ws['B' + str(i)].value is not None:

            stock = ws['C' + str(i)].value.rstrip()
            Stock_Fund = ws['B' + str(i)].value.rstrip()
            # print(ws['F' + str(i)].value)
            quote = get_Current_Stock_Price(stock, Stock_Fund)
            if quote != None:
               ws['G'+ str(i)] = quote
            else:
               print("No data found!", end=" ")
            print(ws['A' + str(i)].value, end="   ")
            print(ws['G'+ str(i)].value)
        i = i + 1            

    wb.save(xcl)

def get_Current_Stock_Price(stock, Stock_Fund):
   today = date.today()

   if Stock_Fund != 'Fund':
       return(float(si.get_live_price(stock)))
   else:
#       data = yf.download(stock, start=today)
#       ticket = yf.Ticker(stock)
       return float(yf.download(stock, start="2022-12-20").iloc[0,4])

def main():
    update_Excel_Table(eXCEL_File)

if __name__ == '__main__':
    
    try:
        request = requests.get("http://www.yahoo.com", timeout=5)
        print("Connected to the Internet")
    except (requests.ConnectionError, requests.Timeout) as exception:
        print("No internet connection.\n\n")
        time.sleep(delay + 3)
        sys.exit()
    
    try:
        wb = load_workbook(eXCEL_File)
    except:
        print("File not found!.... Process Aborted")
        sys.exit()
        
    try:
        wb.save(eXCEL_File)
    except:
        print('Please close the Spreadsheet file, Process Aborted!')
        time.sleep(3)
        sys.exit()
    
    # wb = load_workbook(eXCEL_File)
    # print(wb.sheetnames)
    #ws1 = wb.active
#    std =  wb['Sheet2']
    # wb.remove(wb['Sheet2'])
    # wb.remove(wb['Sheet3'])
    # print(wb.sheetnames)
    # wb.save(eXCEL_File)
    # for sheet in wb.sheetnames:
    #     print(sheet, sheet.title, type(sheet.title))
    #     if str(sheet) == 'Sheet1':
    #         del wb[sheet]
    main()